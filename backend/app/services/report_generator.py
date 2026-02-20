"""Report generation service — JSON, PDF, and Excel export.

Generates professional compliance reports in multiple formats:

- **JSON**: Full ``ComplianceReport`` serialized to JSON.
- **PDF**: WeasyPrint + Jinja2 HTML template → PDF with cover page,
  executive summary, compliance matrix, detailed findings, and appendix.
- **Excel**: openpyxl workbook with Summary, All Rules, and
  Non-Compliant Items sheets, colour-coded by status.
"""

from __future__ import annotations

import io
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from jinja2 import Environment, FileSystemLoader, select_autoescape

logger = logging.getLogger(__name__)

# Template directory — lives alongside the app package
_TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "templates"

# Status → colour mappings for Excel
_STATUS_COLOURS: dict[str, str] = {
    "compliant": "C6EFCE",         # green fill
    "non_compliant": "FFC7CE",     # red fill
    "partially_compliant": "FFEB9C",  # amber fill
    "not_applicable": "D9D9D9",    # grey fill
    "unable_to_determine": "D9D9D9",
}

_STATUS_LABELS: dict[str, str] = {
    "compliant": "Compliant",
    "non_compliant": "Non-Compliant",
    "partially_compliant": "Partially Compliant",
    "not_applicable": "Not Applicable",
    "unable_to_determine": "Unable to Determine",
}


class ReportGenerator:
    """Generates compliance reports in JSON, PDF, and Excel."""

    def __init__(self) -> None:
        _TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
        self._jinja = Environment(
            loader=FileSystemLoader(str(_TEMPLATE_DIR)),
            autoescape=select_autoescape(["html", "xml"]),
        )

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def generate_json(self, report_data: dict[str, Any]) -> bytes:
        """Return the full compliance report as formatted JSON bytes."""
        clean = self._prepare_report_data(report_data)
        return json.dumps(clean, indent=2, default=str).encode("utf-8")

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    def generate_pdf(self, report_data: dict[str, Any]) -> bytes:
        """Generate a professional PDF compliance report using reportlab."""
        context = self._build_template_context(report_data)
        return self._generate_reportlab_pdf(context)

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------

    def generate_excel(self, report_data: dict[str, Any]) -> bytes:
        """Create a multi-sheet Excel workbook and return as bytes."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        except ImportError:
            logger.error("openpyxl is not installed — Excel generation unavailable.")
            raise RuntimeError("openpyxl is not installed")

        wb = Workbook()
        data = self._prepare_report_data(report_data)
        results: list[dict[str, Any]] = data.get("results", [])

        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        def _style_header_row(ws: Any, cols: int) -> None:
            for col_idx in range(1, cols + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = thin_border

        # ── Sheet 1: Summary ──────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary_rows = [
            ("Report ID", data.get("report_id", "")),
            ("Document", data.get("document_name", "")),
            ("Company", data.get("company_name", "N/A")),
            ("Fiscal Year", data.get("fiscal_year", "N/A")),
            ("Frameworks Tested", ", ".join(data.get("frameworks_tested", []))),
            ("Generated At", data.get("generated_at", "")),
            ("Processing Time (s)", data.get("processing_time", 0)),
            ("", ""),
            ("Overall Compliance Score (%)", data.get("overall_compliance_score", 0)),
            ("Total Rules Checked", data.get("total_rules_checked", 0)),
            ("Compliant", data.get("compliant_count", 0)),
            ("Non-Compliant", data.get("non_compliant_count", 0)),
            ("Partially Compliant", data.get("partially_compliant_count", 0)),
            ("Not Applicable", data.get("not_applicable_count", 0)),
            ("Unable to Determine", data.get("unable_to_determine_count", 0)),
            ("", ""),
            ("Executive Summary", data.get("summary", "")),
        ]

        ws_summary.column_dimensions["A"].width = 35
        ws_summary.column_dimensions["B"].width = 80

        for row_idx, (label, value) in enumerate(summary_rows, start=1):
            cell_a = ws_summary.cell(row=row_idx, column=1, value=label)
            cell_a.font = Font(bold=True) if label else Font()
            cell_b = ws_summary.cell(row=row_idx, column=2, value=str(value))
            cell_b.alignment = Alignment(wrap_text=True)

        # ── Sheet 2: All Rules ────────────────────────────────────────
        ws_all = wb.create_sheet("All Rules")
        all_headers = [
            "Rule ID", "Framework", "Rule Source", "Status",
            "Confidence", "Evidence", "Evidence Location",
            "Explanation", "Recommendations", "Rule Text",
        ]
        ws_all.append(all_headers)
        _style_header_row(ws_all, len(all_headers))

        for r in results:
            status_raw = r.get("status", "")
            row = [
                r.get("rule_id", ""),
                r.get("framework", ""),
                r.get("rule_source", ""),
                _STATUS_LABELS.get(status_raw, status_raw),
                r.get("confidence", 0),
                r.get("evidence", ""),
                r.get("evidence_location", ""),
                r.get("explanation", ""),
                r.get("recommendations", ""),
                r.get("rule_text", ""),
            ]
            ws_all.append(row)

            row_num = ws_all.max_row
            status_fill = _STATUS_COLOURS.get(status_raw, "FFFFFF")
            ws_all.cell(row=row_num, column=4).fill = PatternFill(
                start_color=status_fill, end_color=status_fill, fill_type="solid"
            )

        for col_idx, width in enumerate([20, 18, 30, 22, 12, 50, 20, 50, 50, 50], start=1):
            ws_all.column_dimensions[chr(64 + col_idx)].width = width

        # ── Sheet 3: Non-Compliant Items ──────────────────────────────
        ws_nc = wb.create_sheet("Non-Compliant Items")
        nc_headers = [
            "Rule ID", "Framework", "Rule Source", "Status",
            "Confidence", "Evidence", "Explanation", "Recommendations",
        ]
        ws_nc.append(nc_headers)
        _style_header_row(ws_nc, len(nc_headers))

        non_compliant = [
            r for r in results
            if r.get("status") in ("non_compliant", "partially_compliant")
        ]
        for r in non_compliant:
            status_raw = r.get("status", "")
            row = [
                r.get("rule_id", ""),
                r.get("framework", ""),
                r.get("rule_source", ""),
                _STATUS_LABELS.get(status_raw, status_raw),
                r.get("confidence", 0),
                r.get("evidence", ""),
                r.get("explanation", ""),
                r.get("recommendations", ""),
            ]
            ws_nc.append(row)

            row_num = ws_nc.max_row
            status_fill = _STATUS_COLOURS.get(status_raw, "FFFFFF")
            ws_nc.cell(row=row_num, column=4).fill = PatternFill(
                start_color=status_fill, end_color=status_fill, fill_type="solid"
            )

        for col_idx, width in enumerate([20, 18, 30, 22, 12, 50, 50, 50], start=1):
            ws_nc.column_dimensions[chr(64 + col_idx)].width = width

        # Write to bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _prepare_report_data(report_data: dict[str, Any]) -> dict[str, Any]:
        """Normalise a report dict — handles both Pydantic models and raw dicts."""
        if hasattr(report_data, "model_dump"):
            data = report_data.model_dump()
        else:
            data = dict(report_data)

        # Ensure results are plain dicts
        raw_results = data.get("results", [])
        clean_results: list[dict[str, Any]] = []
        for r in raw_results:
            if hasattr(r, "model_dump"):
                clean_results.append(r.model_dump())
            elif isinstance(r, dict):
                clean_results.append(r)
        data["results"] = clean_results
        return data

    def _build_template_context(self, report_data: dict[str, Any]) -> dict[str, Any]:
        """Build a Jinja2-friendly context dict from report data."""
        data = self._prepare_report_data(report_data)
        results = data.get("results", [])

        # Group results by framework
        by_framework: dict[str, list[dict[str, Any]]] = {}
        for r in results:
            fw = r.get("framework", "Unknown")
            by_framework.setdefault(fw, []).append(r)

        # Framework-level stats
        framework_stats: list[dict[str, Any]] = []
        for fw, rules in by_framework.items():
            comp = sum(1 for r in rules if r.get("status") == "compliant")
            nc = sum(1 for r in rules if r.get("status") == "non_compliant")
            pc = sum(1 for r in rules if r.get("status") == "partially_compliant")
            na = sum(1 for r in rules if r.get("status") in ("not_applicable", "unable_to_determine"))
            total = len(rules)
            scorable = comp + nc + pc
            score = (comp + 0.5 * pc) / scorable * 100 if scorable else 0.0
            framework_stats.append({
                "name": fw,
                "total": total,
                "compliant": comp,
                "non_compliant": nc,
                "partially_compliant": pc,
                "not_applicable": na,
                "score": round(score, 1),
            })

        non_compliant_results = [
            r for r in results if r.get("status") == "non_compliant"
        ]
        partial_results = [
            r for r in results if r.get("status") == "partially_compliant"
        ]

        return {
            **data,
            "status_labels": _STATUS_LABELS,
            "status_colours": {
                "compliant": "#30d158",
                "non_compliant": "#ff3b30",
                "partially_compliant": "#ff9f0a",
                "not_applicable": "#86868b",
                "unable_to_determine": "#86868b",
            },
            "framework_stats": framework_stats,
            "non_compliant_results": non_compliant_results,
            "partial_results": partial_results,
            "generated_date": datetime.now(timezone.utc).strftime("%B %d, %Y"),
            "generated_time": datetime.now(timezone.utc).strftime("%H:%M UTC"),
        }

    def _render_html(self, context: dict[str, Any]) -> str:
        """Render the compliance report HTML template."""
        try:
            template = self._jinja.get_template("compliance_report.html")
        except Exception:
            logger.warning("Template not found — using inline fallback")
            return self._fallback_html(context)
        return template.render(**context)

    @staticmethod
    def _fallback_html(ctx: dict[str, Any]) -> str:
        """Minimal inline HTML when the template file is missing."""
        score = ctx.get("overall_compliance_score", 0)
        doc = ctx.get("document_name", "Unknown")
        summary = ctx.get("summary", "")
        return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Compliance Report</title>
<style>body{{font-family:system-ui,sans-serif;margin:40px;}}
h1{{color:#1d1d1f;}}table{{border-collapse:collapse;width:100%;}}
th,td{{border:1px solid #d2d2d7;padding:8px 12px;text-align:left;}}
th{{background:#f5f5f7;}}</style></head><body>
<h1>Compliance Report — {doc}</h1>
<p><strong>Overall Score:</strong> {score:.1f}%</p>
<h2>Executive Summary</h2><p>{summary}</p>
</body></html>"""

    def _generate_reportlab_pdf(self, ctx: dict[str, Any]) -> bytes:
        """Generate PDF using reportlab (pure Python, no system deps)."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, HRFlowable,
        )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=20 * mm, rightMargin=20 * mm,
            topMargin=20 * mm, bottomMargin=20 * mm,
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            "CoverTitle", parent=styles["Title"],
            fontSize=24, spaceAfter=10, textColor=colors.HexColor("#1d1d1f"),
        ))
        styles.add(ParagraphStyle(
            "SectionHead", parent=styles["Heading2"],
            fontSize=14, textColor=colors.HexColor("#0071e3"),
            spaceAfter=8, spaceBefore=16, borderWidth=0,
            borderPadding=0, borderColor=colors.HexColor("#0071e3"),
        ))
        styles.add(ParagraphStyle(
            "BodyText2", parent=styles["BodyText"],
            fontSize=9, leading=13, spaceAfter=6,
        ))
        styles.add(ParagraphStyle(
            "SmallBold", parent=styles["BodyText"],
            fontSize=9, leading=12, fontName="Helvetica-Bold",
        ))

        status_color = {
            "compliant": colors.HexColor("#30d158"),
            "non_compliant": colors.HexColor("#ff3b30"),
            "partially_compliant": colors.HexColor("#ff9f0a"),
            "not_applicable": colors.HexColor("#86868b"),
            "unable_to_determine": colors.HexColor("#86868b"),
        }

        story: list[Any] = []

        # ── Cover ─────────────────────────────────────────────────
        story.append(Spacer(1, 60 * mm))
        story.append(Paragraph("NFRA COMPLIANCE ENGINE", styles["CoverTitle"]))
        story.append(Paragraph("Compliance Validation Report", styles["Heading3"]))
        story.append(Spacer(1, 10 * mm))
        story.append(Paragraph(
            f"<b>Document:</b> {ctx.get('document_name', 'Unknown')}", styles["BodyText2"]
        ))
        if ctx.get("company_name"):
            story.append(Paragraph(
                f"<b>Company:</b> {ctx['company_name']}", styles["BodyText2"]
            ))
        if ctx.get("fiscal_year"):
            story.append(Paragraph(
                f"<b>Fiscal Year:</b> {ctx['fiscal_year']}", styles["BodyText2"]
            ))
        story.append(Paragraph(
            f"<b>Frameworks:</b> {', '.join(ctx.get('frameworks_tested', []))}",
            styles["BodyText2"],
        ))
        story.append(Paragraph(
            f"<b>Generated:</b> {ctx.get('generated_date', '')}",
            styles["BodyText2"],
        ))
        score = ctx.get("overall_compliance_score", 0)
        score_color = "#30d158" if score >= 80 else "#ff9f0a" if score >= 50 else "#ff3b30"
        story.append(Spacer(1, 10 * mm))
        story.append(Paragraph(
            f'<font size="20" color="{score_color}"><b>{score:.1f}%</b></font> '
            f'<font size="12">Overall Compliance</font>',
            styles["BodyText2"],
        ))
        story.append(PageBreak())

        # ── Executive Summary ────────────────────────────────────
        story.append(Paragraph("Executive Summary", styles["SectionHead"]))
        story.append(HRFlowable(
            width="100%", thickness=1, color=colors.HexColor("#0071e3"),
        ))
        story.append(Spacer(1, 4 * mm))

        summary_data = [
            ["Rules Checked", str(ctx.get("total_rules_checked", 0))],
            ["Compliant", str(ctx.get("compliant_count", 0))],
            ["Non-Compliant", str(ctx.get("non_compliant_count", 0))],
            ["Partially Compliant", str(ctx.get("partially_compliant_count", 0))],
            ["Not Applicable", str(ctx.get("not_applicable_count", 0))],
        ]
        t = Table(summary_data, colWidths=[100 * mm, 60 * mm])
        t.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e5e7")),
        ]))
        story.append(t)
        story.append(Spacer(1, 6 * mm))

        summary_text = ctx.get("summary", "")
        if summary_text:
            for para in summary_text.split("\n"):
                if para.strip():
                    story.append(Paragraph(para.strip(), styles["BodyText2"]))

        # ── Framework Scores ─────────────────────────────────────
        fw_stats = ctx.get("framework_stats", [])
        if fw_stats:
            story.append(Spacer(1, 8 * mm))
            story.append(Paragraph("Score by Framework", styles["SectionHead"]))
            fw_header = ["Framework", "Total", "Compliant", "Non-Compliant", "Partial", "Score"]
            fw_rows = [fw_header]
            for fw in fw_stats:
                fw_rows.append([
                    fw["name"], str(fw["total"]), str(fw["compliant"]),
                    str(fw["non_compliant"]), str(fw["partially_compliant"]),
                    f'{fw["score"]}%',
                ])
            ft = Table(fw_rows, colWidths=[50 * mm, 18 * mm, 24 * mm, 28 * mm, 18 * mm, 18 * mm])
            ft.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f7")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d2d2d7")),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ]))
            story.append(ft)

        story.append(PageBreak())

        # ── Compliance Matrix ────────────────────────────────────
        results = ctx.get("results", [])
        if results:
            story.append(Paragraph("Compliance Matrix", styles["SectionHead"]))
            matrix_header = ["Rule Source", "Framework", "Status", "Confidence"]
            matrix_rows = [matrix_header]
            for r in results[:100]:
                status = r.get("status", "")
                label = _STATUS_LABELS.get(status, status)
                conf = r.get("confidence", 0)
                matrix_rows.append([
                    Paragraph(str(r.get("rule_source", ""))[:60], styles["BodyText2"]),
                    str(r.get("framework", "")),
                    label,
                    f"{conf * 100:.0f}%",
                ])

            mt = Table(matrix_rows, colWidths=[60 * mm, 30 * mm, 40 * mm, 26 * mm])
            row_styles = [
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f7")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d2d2d7")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
            for i, r in enumerate(results[:100], start=1):
                bg = status_color.get(r.get("status", ""), colors.white)
                row_styles.append(("BACKGROUND", (2, i), (2, i), bg))
                row_styles.append(("TEXTCOLOR", (2, i), (2, i), colors.white))
            mt.setStyle(TableStyle(row_styles))
            story.append(mt)
            story.append(PageBreak())

        # ── Non-Compliant Findings ───────────────────────────────
        nc_results = ctx.get("non_compliant_results", [])
        if nc_results:
            story.append(Paragraph("Non-Compliant Findings", styles["SectionHead"]))
            for r in nc_results[:30]:
                story.append(Paragraph(
                    f"<b>{r.get('rule_source', 'Unknown')}</b> — {r.get('framework', '')}",
                    styles["SmallBold"],
                ))
                if r.get("rule_text"):
                    story.append(Paragraph(
                        f"<i>Requirement:</i> {str(r['rule_text'])[:300]}", styles["BodyText2"]
                    ))
                if r.get("evidence"):
                    story.append(Paragraph(
                        f"<i>Evidence:</i> {str(r['evidence'])[:300]}", styles["BodyText2"]
                    ))
                if r.get("explanation"):
                    story.append(Paragraph(
                        f"<i>Explanation:</i> {str(r['explanation'])[:400]}", styles["BodyText2"]
                    ))
                if r.get("recommendations"):
                    story.append(Paragraph(
                        f"<i>Recommendations:</i> {str(r['recommendations'])[:300]}",
                        styles["BodyText2"],
                    ))
                story.append(Spacer(1, 4 * mm))
                story.append(HRFlowable(
                    width="100%", thickness=0.5, color=colors.HexColor("#e5e5e7"),
                ))
                story.append(Spacer(1, 2 * mm))

        # ── Footer / Appendix ────────────────────────────────────
        story.append(PageBreak())
        story.append(Paragraph("Appendix", styles["SectionHead"]))
        story.append(Paragraph(
            f"<b>Report ID:</b> {ctx.get('report_id', '')}", styles["BodyText2"]
        ))
        story.append(Paragraph(
            f"<b>Document ID:</b> {ctx.get('document_id', '')}", styles["BodyText2"]
        ))
        story.append(Paragraph(
            f"<b>Processing Time:</b> {ctx.get('processing_time', 0):.1f}s",
            styles["BodyText2"],
        ))
        story.append(Spacer(1, 20 * mm))
        story.append(Paragraph(
            f"Generated by NFRA Compliance Engine — {ctx.get('generated_date', '')}",
            ParagraphStyle("Footer", parent=styles["BodyText"], fontSize=8,
                           textColor=colors.HexColor("#86868b"), alignment=1),
        ))

        doc.build(story)
        buffer.seek(0)
        return buffer.read()
